"""
ui/pages/write_off.py — Акт списания расходников
=================================================

Форма для оформления списания запчастей на конкретный автомобиль
с привязкой к водителю и экспортом в Excel.
Акты хранятся в БД, группируются по водителям/авто.
"""

from __future__ import annotations

import sqlite3
from datetime import datetime
from typing import List, Optional

from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton,
    QTableWidget, QTableWidgetItem, QHeaderView, QDialog,
    QFormLayout, QComboBox, QDoubleSpinBox, QTextEdit,
    QMessageBox, QDateEdit, QFileDialog, QAbstractItemView,
    QTreeWidget, QTreeWidgetItem, QMenu,
)
from PySide6.QtCore import Qt, QDate
from PySide6.QtGui import QColor, QAction

from zavgar_app import db
from zavgar_app.models import Part, PartTransaction, Vehicle, Driver


def _ensure_tables(conn: sqlite3.Connection):
    """Создать таблицы актов списания если не существуют."""
    conn.execute("""
        CREATE TABLE IF NOT EXISTS write_offs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            act_number TEXT NOT NULL,
            act_date TEXT NOT NULL,
            vehicle_id INTEGER,
            driver_id INTEGER,
            reason TEXT,
            total_amount REAL DEFAULT 0,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (vehicle_id) REFERENCES vehicles(id),
            FOREIGN KEY (driver_id) REFERENCES drivers(id)
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS write_off_items (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            write_off_id INTEGER NOT NULL,
            part_id INTEGER,
            part_name TEXT,
            article TEXT,
            unit TEXT DEFAULT 'шт',
            quantity REAL NOT NULL,
            price REAL DEFAULT 0,
            amount REAL DEFAULT 0,
            FOREIGN KEY (write_off_id) REFERENCES write_offs(id)
        )
    """)
    conn.commit()


def _next_act_number(conn: sqlite3.Connection) -> str:
    """Сгенерировать следующий номер акта."""
    try:
        cur = conn.execute("SELECT MAX(id) FROM write_offs")
        row = cur.fetchone()
        last_id = row[0] if row[0] else 0
    except Exception:
        last_id = 0
    return f"АС-{last_id + 1:04d}"


class WriteOffDialog(QDialog):
    """Диалог создания акта списания."""

    def __init__(self, conn: sqlite3.Connection, parent=None):
        super().__init__(parent)
        self.conn = conn
        self.setWindowTitle("Акт списания расходников")
        self.setMinimumWidth(700)
        self.setMinimumHeight(500)
        
        self.items: List[dict] = []
        self.saved_act_id: Optional[int] = None
        
        self._setup_ui()
        self._load_data()

    def _setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setSpacing(16)

        title = QLabel("📋 Акт списания расходных материалов")
        title.setStyleSheet("font-size: 18px; font-weight: 700;")
        layout.addWidget(title)

        form = QFormLayout()
        form.setSpacing(10)
        
        # Номер акта
        self.act_number = _next_act_number(self.conn)
        num_label = QLabel(self.act_number)
        num_label.setStyleSheet("font-weight: 700; color: #6366f1; font-size: 14px;")
        form.addRow("Номер акта:", num_label)

        self.date_edit = QDateEdit()
        self.date_edit.setCalendarPopup(True)
        self.date_edit.setDate(QDate.currentDate())
        form.addRow("Дата:", self.date_edit)

        self.vehicle_combo = QComboBox()
        self.vehicle_combo.setEditable(True)
        self.vehicle_combo.setInsertPolicy(QComboBox.NoInsert)
        form.addRow("Автомобиль:", self.vehicle_combo)

        self.driver_combo = QComboBox()
        self.driver_combo.setEditable(True)
        self.driver_combo.setInsertPolicy(QComboBox.NoInsert)
        form.addRow("Водитель:", self.driver_combo)

        self.reason_edit = QTextEdit()
        self.reason_edit.setPlaceholderText("Плановое ТО, замена изношенных деталей...")
        self.reason_edit.setMaximumHeight(60)
        # Стили из темы, не inline
        form.addRow("Причина:", self.reason_edit)

        layout.addLayout(form)

        items_label = QLabel("Позиции для списания:")
        items_label.setStyleSheet("font-weight: 600;")
        layout.addWidget(items_label)

        self.items_table = QTableWidget()
        self.items_table.setColumnCount(5)
        self.items_table.setHorizontalHeaderLabels([
            "Запчасть", "Артикул", "Количество", "Цена", "Сумма"
        ])
        self.items_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)
        self.items_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeToContents)
        self.items_table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeToContents)
        self.items_table.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeToContents)
        self.items_table.horizontalHeader().setSectionResizeMode(4, QHeaderView.ResizeToContents)
        self.items_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.items_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.items_table.verticalHeader().setVisible(False)
        self.items_table.setAlternatingRowColors(True)
        # Стили из темы, не inline
        layout.addWidget(self.items_table)

        items_btn_layout = QHBoxLayout()
        
        add_item_btn = QPushButton("➕ Добавить позицию")
        add_item_btn.setObjectName("primaryBtn")
        add_item_btn.clicked.connect(self._add_item)
        items_btn_layout.addWidget(add_item_btn)

        remove_item_btn = QPushButton("🗑️ Удалить позицию")
        remove_item_btn.setObjectName("dangerBtn")
        remove_item_btn.clicked.connect(self._remove_item)
        items_btn_layout.addWidget(remove_item_btn)

        items_btn_layout.addStretch()
        layout.addLayout(items_btn_layout)

        self.total_label = QLabel("Итого: ₸ 0.00")
        self.total_label.setStyleSheet("font-size: 16px; font-weight: 700; color: #10b981;")
        self.total_label.setAlignment(Qt.AlignRight)
        layout.addWidget(self.total_label)

        btn_layout = QHBoxLayout()
        btn_layout.addStretch()

        cancel_btn = QPushButton("Отмена")
        cancel_btn.clicked.connect(self.reject)
        btn_layout.addWidget(cancel_btn)

        export_btn = QPushButton("📊 Экспорт в Excel")
        export_btn.clicked.connect(self._export_to_excel)
        btn_layout.addWidget(export_btn)

        save_btn = QPushButton("💾 Сохранить акт")
        save_btn.setObjectName("primaryBtn")
        save_btn.clicked.connect(self._save_write_off)
        btn_layout.addWidget(save_btn)

        layout.addLayout(btn_layout)

    def _load_data(self):
        vehicles = db.list_vehicles(self.conn)
        self.vehicle_combo.addItem("— Выберите автомобиль —", None)
        for v in vehicles:
            self.vehicle_combo.addItem(f"{v.marka} {v.model} ({v.gosnomer})", v.id)

        drivers = db.list_drivers(self.conn)
        self.driver_combo.addItem("— Выберите водителя —", None)
        for d in drivers:
            self.driver_combo.addItem(d.fio, d.id)

    def _add_item(self):
        dialog = WriteOffItemDialog(self.conn, parent=self)
        if dialog.exec() == QDialog.Accepted:
            item = dialog.get_item()
            self.items.append(item)
            self._refresh_items_table()

    def _remove_item(self):
        rows = self.items_table.selectionModel().selectedRows()
        if not rows:
            QMessageBox.information(self, "Подсказка", "Выберите позицию для удаления")
            return
        row = rows[0].row()
        if 0 <= row < len(self.items):
            del self.items[row]
            self._refresh_items_table()

    def _refresh_items_table(self):
        self.items_table.setRowCount(len(self.items))
        total = 0.0

        for row, item in enumerate(self.items):
            for col, val in enumerate([
                item["name"], item["article"],
                f"{item['quantity']:.2f} {item['unit']}",
                f"₸ {item['price']:,.2f}",
            ]):
                cell = QTableWidgetItem(val)
                cell.setForeground(QColor('#e4e4e7'))
                self.items_table.setItem(row, col, cell)

            amount = item['quantity'] * item['price']
            total += amount
            amount_item = QTableWidgetItem(f"₸ {amount:,.2f}")
            amount_item.setForeground(QColor('#10b981'))
            self.items_table.setItem(row, 4, amount_item)

        self.total_label.setText(f"Итого: ₸ {total:,.2f}")

    def _save_write_off(self):
        vehicle_id = self.vehicle_combo.currentData()
        if not vehicle_id:
            QMessageBox.warning(self, "Ошибка", "Выберите автомобиль")
            return

        driver_id = self.driver_combo.currentData()
        if not driver_id:
            QMessageBox.warning(self, "Ошибка", "Выберите водителя")
            return

        if not self.items:
            QMessageBox.warning(self, "Ошибка", "Добавьте хотя бы одну позицию")
            return

        _ensure_tables(self.conn)

        date_str = self.date_edit.date().toString("yyyy-MM-dd")
        reason = self.reason_edit.toPlainText().strip()
        total = sum(i['quantity'] * i['price'] for i in self.items)

        try:
            # 1. Сохранить акт
            cur = self.conn.execute(
                """INSERT INTO write_offs (act_number, act_date, vehicle_id, driver_id, reason, total_amount)
                   VALUES (?, ?, ?, ?, ?, ?)""",
                (self.act_number, date_str, vehicle_id, driver_id, reason, total)
            )
            act_id = cur.lastrowid

            # 2. Сохранить позиции + списать остатки
            for item in self.items:
                part_id = item["part_id"]
                quantity = item["quantity"]
                price = item["price"]
                amount = quantity * price

                self.conn.execute(
                    """INSERT INTO write_off_items (write_off_id, part_id, part_name, article, unit, quantity, price, amount)
                       VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
                    (act_id, part_id, item["name"], item["article"], item["unit"], quantity, price, amount)
                )

                # Транзакция расхода
                tx = PartTransaction(
                    part_id=part_id,
                    transaction_type='out',
                    quantity=quantity,
                    price=price,
                    notes=f"Акт {self.act_number}: {reason}",
                    transaction_date=date_str,
                )
                db.create_part_transaction(self.conn, tx)

                # Обновить остаток
                part = db.get_part(self.conn, part_id)
                if part:
                    new_qty = max(0, part.quantity - quantity)
                    db.update_part_quantity(self.conn, part_id, new_qty)

            self.conn.commit()
            self.saved_act_id = act_id
            QMessageBox.information(self, "Успех", f"Акт {self.act_number} сохранён")
            self.accept()

        except Exception as e:
            self.conn.rollback()
            QMessageBox.critical(self, "Ошибка", f"Не удалось сохранить: {e}")

    def _export_to_excel(self):
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
        except ImportError:
            QMessageBox.warning(self, "Ошибка", "openpyxl не установлен: uv pip install openpyxl")
            return

        if not self.items:
            QMessageBox.warning(self, "Ошибка", "Нет позиций для экспорта")
            return

        file_path, _ = QFileDialog.getSaveFileName(
            self, "Сохранить акт списания",
            f"Акт_{self.act_number}_{datetime.now().strftime('%Y%m%d')}.xlsx",
            "Excel Files (*.xlsx)"
        )
        if not file_path:
            return

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Акт списания"

            ws.merge_cells('A1:F1')
            c = ws['A1']
            c.value = f"АКТ СПИСАНИЯ № {self.act_number}"
            c.font = Font(size=14, bold=True)
            c.alignment = Alignment(horizontal='center')

            ws['A3'] = "Дата:"
            ws['B3'] = self.date_edit.date().toString("dd.MM.yyyy")
            ws['A4'] = "Автомобиль:"
            ws['B4'] = self.vehicle_combo.currentText()
            ws['A5'] = "Водитель:"
            ws['B5'] = self.driver_combo.currentText()
            ws['A6'] = "Причина:"
            ws['B6'] = self.reason_edit.toPlainText().strip()

            row = 8
            headers = ["№", "Наименование", "Артикул", "Кол-во", "Цена", "Сумма"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="E5E7EB", end_color="E5E7EB", fill_type="solid")
                cell.alignment = Alignment(horizontal='center')

            total = 0.0
            for idx, item in enumerate(self.items, 1):
                row += 1
                ws.cell(row=row, column=1, value=idx)
                ws.cell(row=row, column=2, value=item["name"])
                ws.cell(row=row, column=3, value=item["article"])
                ws.cell(row=row, column=4, value=f"{item['quantity']:.2f} {item['unit']}")
                ws.cell(row=row, column=5, value=item['price'])
                amount = item['quantity'] * item['price']
                total += amount
                ws.cell(row=row, column=6, value=amount)

            row += 1
            ws.merge_cells(f'A{row}:E{row}')
            tc = ws.cell(row=row, column=1, value="ИТОГО:")
            tc.font = Font(bold=True)
            tc.alignment = Alignment(horizontal='right')
            ws.cell(row=row, column=6, value=total).font = Font(bold=True)

            row += 3
            ws.cell(row=row, column=1, value="Списал: ____________________")
            ws.cell(row=row, column=4, value="Принял: ____________________")

            ws.column_dimensions['A'].width = 5
            ws.column_dimensions['B'].width = 30
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 15
            ws.column_dimensions['E'].width = 12
            ws.column_dimensions['F'].width = 15

            wb.save(file_path)
            QMessageBox.information(self, "Успех", f"Файл сохранён: {file_path}")

        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Не удалось экспортировать: {e}")


class WriteOffItemDialog(QDialog):
    """Диалог добавления позиции в акт списания."""

    def __init__(self, conn: sqlite3.Connection, parent=None):
        super().__init__(parent)
        self.conn = conn
        self.setWindowTitle("Добавить позицию")
        self.setMinimumWidth(400)

        layout = QVBoxLayout(self)
        form = QFormLayout()
        form.setSpacing(10)

        self.part_combo = QComboBox()
        self.part_combo.setEditable(True)
        self.part_combo.setInsertPolicy(QComboBox.NoInsert)
        parts = db.list_parts(self.conn)
        for p in parts:
            self.part_combo.addItem(f"{p.name} (остаток: {p.quantity} {p.unit})", p.id)
        self.part_combo.currentIndexChanged.connect(self._on_part_changed)
        form.addRow("Запчасть:", self.part_combo)

        self.quantity_spin = QDoubleSpinBox()
        self.quantity_spin.setRange(0.01, 99_999)
        self.quantity_spin.setDecimals(2)
        form.addRow("Количество:", self.quantity_spin)

        self.price_spin = QDoubleSpinBox()
        self.price_spin.setRange(0, 999_999)
        self.price_spin.setDecimals(2)
        self.price_spin.setPrefix("₸ ")
        form.addRow("Цена:", self.price_spin)

        layout.addLayout(form)

        btn_layout = QHBoxLayout()
        btn_layout.addStretch()
        cancel_btn = QPushButton("Отмена")
        cancel_btn.clicked.connect(self.reject)
        btn_layout.addWidget(cancel_btn)
        save_btn = QPushButton("Добавить")
        save_btn.setObjectName("primaryBtn")
        save_btn.clicked.connect(self.accept)
        btn_layout.addWidget(save_btn)
        layout.addLayout(btn_layout)

        # Автозаполнение цены
        self._on_part_changed()

    def _on_part_changed(self):
        part_id = self.part_combo.currentData()
        if not part_id:
            return
        part = db.get_part(self.conn, part_id)
        if part and part.avg_price:
            self.price_spin.setValue(part.avg_price)

    def get_item(self) -> dict:
        part_id = self.part_combo.currentData()
        part = db.get_part(self.conn, part_id) if part_id else None
        name = part.name if part else self.part_combo.currentText().strip()
        if not name:
            name = "Неизвестная запчасть"
        return {
            "part_id": part_id or 0,
            "name": name,
            "article": part.article if part else "",
            "quantity": self.quantity_spin.value(),
            "unit": part.unit if part else "шт",
            "price": self.price_spin.value(),
        }


class WriteOffPage(QWidget):
    """Страница актов списания с группировкой по водителям/авто."""

    def __init__(self, conn: sqlite3.Connection, parent=None):
        super().__init__(parent)
        self.conn = conn
        _ensure_tables(self.conn)
        self._setup_ui()

    def _setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(32, 24, 32, 24)
        layout.setSpacing(20)

        # Заголовок
        header = QHBoxLayout()
        title = QLabel("📋 Акты списания")
        title.setObjectName("pageTitle")
        header.addWidget(title)
        header.addStretch()

        add_btn = QPushButton("➕ Создать акт")
        add_btn.setObjectName("primaryBtn")
        add_btn.clicked.connect(self._create_write_off)
        header.addWidget(add_btn)

        export_all_btn = QPushButton("📊 Экспорт всех")
        export_all_btn.clicked.connect(self._export_all)
        header.addWidget(export_all_btn)

        layout.addLayout(header)

        # Фильтр группировки
        filter_layout = QHBoxLayout()
        filter_layout.addWidget(QLabel("Группировка:"))
        
        self.group_combo = QComboBox()
        self.group_combo.addItems(["По водителям", "По автомобилям", "По датам"])
        self.group_combo.currentIndexChanged.connect(self._refresh)
        filter_layout.addWidget(self.group_combo)
        
        filter_layout.addStretch()
        layout.addLayout(filter_layout)

        # Toolbar действий
        toolbar = QHBoxLayout()
        edit_btn = QPushButton("✏️")
        edit_btn.setObjectName("actionBtn")
        edit_btn.setToolTip("Редактировать акт")
        edit_btn.clicked.connect(self._edit_selected)
        toolbar.addWidget(edit_btn)

        del_btn = QPushButton("🗑️")
        del_btn.setObjectName("actionDelete")
        del_btn.setToolTip("Удалить акт")
        del_btn.clicked.connect(self._delete_selected)
        toolbar.addWidget(del_btn)

        export_btn = QPushButton("🖨️")
        export_btn.setObjectName("ghostBtn")
        export_btn.setToolTip("Печать/Экспорт")
        export_btn.clicked.connect(self._export_selected)
        toolbar.addWidget(export_btn)

        toolbar.addStretch()
        layout.addLayout(toolbar)

        # Дерево актов
        self.tree = QTreeWidget()
        self.tree.setHeaderLabels(["Акт / Позиция", "Дата", "Авто", "Водитель", "Сумма"])
        self.tree.setColumnCount(5)
        self.tree.header().setSectionResizeMode(0, QHeaderView.Stretch)
        self.tree.header().setSectionResizeMode(1, QHeaderView.Fixed)
        self.tree.setColumnWidth(1, 100)
        self.tree.header().setSectionResizeMode(2, QHeaderView.Fixed)
        self.tree.setColumnWidth(2, 150)
        self.tree.header().setSectionResizeMode(3, QHeaderView.Fixed)
        self.tree.setColumnWidth(3, 150)
        self.tree.header().setSectionResizeMode(4, QHeaderView.Fixed)
        self.tree.setColumnWidth(4, 120)
        self.tree.setAlternatingRowColors(True)
        self.tree.setContextMenuPolicy(Qt.CustomContextMenu)
        self.tree.customContextMenuRequested.connect(self._show_context_menu)
        self.tree.doubleClicked.connect(self._on_double_click)
        layout.addWidget(self.tree)



        self._refresh()

    def _refresh(self):
        self.tree.clear()
        group_mode = self.group_combo.currentIndex()  # 0=водители, 1=авто, 2=даты

        # Получить все акты
        acts = self._load_acts()
        if not acts:
            return

        # Получить имена
        drivers = {d.id: d.fio for d in db.list_drivers(self.conn)}
        vehicles = {v.id: f"{v.marka} {v.model} ({v.gosnomer})" for v in db.list_vehicles(self.conn)}

        if group_mode == 0:  # По водителям
            groups = {}
            for act in acts:
                key = drivers.get(act['driver_id'], 'Неизвестный')
                groups.setdefault(key, []).append(act)
            for group_name, group_acts in sorted(groups.items()):
                parent = QTreeWidgetItem([f"👤 {group_name} ({len(group_acts)} актов)", "", "", "", ""])
                parent.setExpanded(True)
                self.tree.addTopLevelItem(parent)
                for act in group_acts:
                    self._add_act_item(parent, act, vehicles, drivers)

        elif group_mode == 1:  # По авто
            groups = {}
            for act in acts:
                key = vehicles.get(act['vehicle_id'], 'Неизвестное авто')
                groups.setdefault(key, []).append(act)
            for group_name, group_acts in sorted(groups.items()):
                parent = QTreeWidgetItem([f"🚗 {group_name} ({len(group_acts)} актов)", "", "", "", ""])
                parent.setExpanded(True)
                self.tree.addTopLevelItem(parent)
                for act in group_acts:
                    self._add_act_item(parent, act, vehicles, drivers)

        else:  # По датам
            groups = {}
            for act in acts:
                key = act['act_date'][:7]  # YYYY-MM
                groups.setdefault(key, []).append(act)
            for group_name, group_acts in sorted(groups.items(), reverse=True):
                parent = QTreeWidgetItem([f"📅 {group_name} ({len(group_acts)} актов)", "", "", "", ""])
                parent.setExpanded(True)
                self.tree.addTopLevelItem(parent)
                for act in group_acts:
                    self._add_act_item(parent, act, vehicles, drivers)

    def _add_act_item(self, parent, act, vehicles, drivers):
        vehicle_name = vehicles.get(act['vehicle_id'], '?')
        driver_name = drivers.get(act['driver_id'], '?')
        
        item = QTreeWidgetItem([
            f"📋 {act['act_number']}",
            act['act_date'],
            vehicle_name,
            driver_name,
            f"₸ {act['total_amount']:,.2f}"
        ])
        item.setData(0, Qt.UserRole, act['id'])
        parent.addChild(item)

        # Позиции
        items = self._load_act_items(act['id'])
        for wo_item in items:
            child = QTreeWidgetItem([
                f"  • {wo_item['part_name']}",
                "",
                wo_item['article'] or '',
                f"{wo_item['quantity']:.2f} {wo_item['unit']}",
                f"₸ {wo_item['amount']:,.2f}"
            ])
            child.setForeground(0, QColor('#9ca3af'))
            item.addChild(child)

    def _load_acts(self) -> list:
        try:
            cur = self.conn.execute(
                "SELECT id, act_number, act_date, vehicle_id, driver_id, reason, total_amount "
                "FROM write_offs ORDER BY act_date DESC"
            )
            return [dict(zip(
                ['id', 'act_number', 'act_date', 'vehicle_id', 'driver_id', 'reason', 'total_amount'],
                row
            )) for row in cur.fetchall()]
        except Exception:
            return []

    def _load_act_items(self, act_id: int) -> list:
        try:
            cur = self.conn.execute(
                "SELECT part_name, article, unit, quantity, price, amount "
                "FROM write_off_items WHERE write_off_id = ?", (act_id,)
            )
            return [dict(zip(
                ['part_name', 'article', 'unit', 'quantity', 'price', 'amount'],
                row
            )) for row in cur.fetchall()]
        except Exception:
            return []

    def _create_write_off(self):
        dialog = WriteOffDialog(self.conn, parent=self)
        if dialog.exec() == QDialog.Accepted:
            self._refresh()

    def _get_selected_act_id(self) -> Optional[int]:
        item = self.tree.currentItem()
        while item:
            act_id = item.data(0, Qt.UserRole)
            if act_id:
                return act_id
            item = item.parent()
        return None

    def _show_context_menu(self, pos):
        """Контекстное меню по правому клику."""
        item = self.tree.itemAt(pos)
        if not item:
            return
        act_id = self._get_selected_act_id()
        if not act_id:
            return
        menu = QMenu(self)
        menu.addAction("✏️ Редактировать", self._edit_selected)
        menu.addAction("🗑️ Удалить", self._delete_selected)
        menu.addSeparator()
        menu.addAction("🖨️ Печать/Экспорт", self._export_selected)
        menu.exec(self.tree.viewport().mapToGlobal(pos))

    def _edit_selected(self):
        """Редактировать выбранный акт."""
        act_id = self._get_selected_act_id()
        if act_id:
            self._show_act_dialog(act_id)
        else:
            QMessageBox.information(self, "Подсказка", "Выберите акт для редактирования")

    def _on_double_click(self, index):
        act_id = self._get_selected_act_id()
        if act_id:
            self._show_act_dialog(act_id)

    def _show_act_dialog(self, act_id: int):
        """Показать акт списания в диалоге редактирования с позициями."""
        acts = self._load_acts()
        act = next((a for a in acts if a['id'] == act_id), None)
        if not act:
            return

        drivers_list = db.list_drivers(self.conn)
        vehicles_list = db.list_vehicles(self.conn)
        drivers = {d.id: d.fio for d in drivers_list}
        vehicles = {v.id: f"{v.marka} {v.model} ({v.gosnomer})" for v in vehicles_list}
        items = self._load_act_items(act_id)

        dialog = QDialog(self)
        dialog.setWindowTitle(f"Акт {act['act_number']}")
        dialog.setMinimumWidth(650)
        dialog.setMinimumHeight(500)
        layout = QVBoxLayout(dialog)

        # Заголовок
        title = QLabel(f"📋 Акт {act['act_number']}")
        title.setStyleSheet("font-size: 18px; font-weight: 700;")
        layout.addWidget(title)

        # Форма редактирования
        form = QFormLayout()
        form.setSpacing(10)

        date_edit = QDateEdit()
        date_edit.setCalendarPopup(True)
        from PySide6.QtCore import QDate as QD
        date_edit.setDate(QD.fromString(act['act_date'], "yyyy-MM-dd"))
        form.addRow("Дата:", date_edit)

        vehicle_combo = QComboBox()
        vehicle_combo.addItem("— Не выбрано —", None)
        for v in vehicles_list:
            vehicle_combo.addItem(f"{v.marka} {v.model} ({v.gosnomer})", v.id)
        idx = vehicle_combo.findData(act['vehicle_id'])
        if idx >= 0:
            vehicle_combo.setCurrentIndex(idx)
        form.addRow("Автомобиль:", vehicle_combo)

        driver_combo = QComboBox()
        driver_combo.addItem("— Не выбрано —", None)
        for d in drivers_list:
            driver_combo.addItem(d.fio, d.id)
        idx = driver_combo.findData(act['driver_id'])
        if idx >= 0:
            driver_combo.setCurrentIndex(idx)
        form.addRow("Водитель:", driver_combo)

        reason_edit = QTextEdit()
        reason_edit.setPlainText(act['reason'] or '')
        reason_edit.setMaximumHeight(60)
        form.addRow("Причина:", reason_edit)

        layout.addLayout(form)

        # Таблица позиций
        items_label = QLabel("Позиции:")
        items_label.setStyleSheet("font-weight: 600;")
        layout.addWidget(items_label)

        table = QTableWidget()
        table.setColumnCount(5)
        table.setHorizontalHeaderLabels(["Запчасть", "Артикул", "Кол-во", "Цена", "Сумма"])
        table.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)
        for col in range(1, 5):
            table.horizontalHeader().setSectionResizeMode(col, QHeaderView.ResizeToContents)
        table.verticalHeader().setVisible(False)
        table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        table.setSelectionBehavior(QAbstractItemView.SelectRows)
        table.setAlternatingRowColors(True)

        for item in items:
            row = table.rowCount()
            table.insertRow(row)
            table.setItem(row, 0, QTableWidgetItem(item['part_name'] or '—'))
            table.setItem(row, 1, QTableWidgetItem(item['article'] or '—'))
            table.setItem(row, 2, QTableWidgetItem(f"{item['quantity']:.2f} {item['unit']}"))
            table.setItem(row, 3, QTableWidgetItem(f"₸ {item['price']:,.2f}"))
            table.setItem(row, 4, QTableWidgetItem(f"₸ {item['amount']:,.2f}"))

        layout.addWidget(table)

        # Итого
        total_amount = sum(i['amount'] for i in items)
        total = QLabel(f"Итого: ₸ {total_amount:,.2f}")
        total.setStyleSheet("font-size: 16px; font-weight: 700;")
        layout.addWidget(total)

        # Кнопки
        btn_layout = QHBoxLayout()
        save_btn = QPushButton("💾 Сохранить")
        save_btn.setObjectName("primaryBtn")

        def _save_act():
            self.conn.execute("""
                UPDATE write_offs SET act_date=?, vehicle_id=?, driver_id=?, reason=?
                WHERE id=?
            """, (
                date_edit.date().toString("yyyy-MM-dd"),
                vehicle_combo.currentData(),
                driver_combo.currentData(),
                reason_edit.toPlainText().strip(),
                act_id,
            ))
            self.conn.commit()
            self._refresh()
            dialog.accept()

        save_btn.clicked.connect(_save_act)
        btn_layout.addWidget(save_btn)

        print_btn = QPushButton("🖨️ Печать")
        print_btn.clicked.connect(lambda: self._print_act(act_id))
        btn_layout.addWidget(print_btn)

        export_btn = QPushButton("📊 Экспорт")
        export_btn.clicked.connect(lambda: self._export_act(act_id))
        btn_layout.addWidget(export_btn)

        btn_layout.addStretch()

        close_btn = QPushButton("Закрыть")
        close_btn.clicked.connect(dialog.reject)
        btn_layout.addWidget(close_btn)

        layout.addLayout(btn_layout)
        dialog.exec()

    def _print_act(self, act_id: int):
        """Печать акта списания через QPrinter."""
        from zavgar_app.utils.print_utils import print_document
        acts = self._load_acts()
        act = next((a for a in acts if a['id'] == act_id), None)
        if not act:
            return

        drivers = {d.id: d.fio for d in db.list_drivers(self.conn)}
        vehicles = {v.id: f"{v.marka} {v.model} ({v.gosnomer})" for v in db.list_vehicles(self.conn)}
        items = self._load_act_items(act_id)

        rows = []
        for item in items:
            rows.append([
                item['part_name'] or '—',
                item['article'] or '—',
                f"{item['quantity']:.2f} {item['unit']}",
                f"₸ {item['price']:,.2f}",
                f"₸ {item['amount']:,.2f}",
            ])

        title = (f"Акт списания {act['act_number']}\n\n"
                 f"Дата: {act['act_date']}\n"
                 f"Авто: {vehicles.get(act['vehicle_id'], '—')}\n"
                 f"Водитель: {drivers.get(act['driver_id'], '—')}\n"
                 f"Причина: {act['reason'] or '—'}\n\n"
                 f"Итого: ₸ {act['total_amount']:,.2f}")

        print_document(title, ["Запчасть", "Артикул", "Кол-во", "Цена", "Сумма"], rows, self)

    def _export_act(self, act_id: int):
        """Экспортировать один акт в Excel."""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
        except ImportError:
            QMessageBox.warning(self, "Ошибка", "openpyxl не установлен")
            return

        acts = self._load_acts()
        act = next((a for a in acts if a['id'] == act_id), None)
        if not act:
            return
        items = self._load_act_items(act_id)

        file_path, _ = QFileDialog.getSaveFileName(
            self, "Экспорт акта", f"{act['act_number']}.xlsx", "Excel (*.xlsx)"
        )
        if not file_path:
            return

        drivers = {d.id: d.fio for d in db.list_drivers(self.conn)}
        vehicles = {v.id: f"{v.marka} {v.model} ({v.gosnomer})" for v in db.list_vehicles(self.conn)}

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Акт списания"

        ws.merge_cells('A1:E1')
        c = ws['A1']
        c.value = f"АКТ СПИСАНИЯ № {act['act_number']}"
        c.font = Font(size=14, bold=True)
        c.alignment = Alignment(horizontal='center')

        ws['A3'] = "Дата:"
        ws['B3'] = act['act_date']
        ws['A4'] = "Автомобиль:"
        ws['B4'] = vehicles.get(act['vehicle_id'], '—')
        ws['A5'] = "Водитель:"
        ws['B5'] = drivers.get(act['driver_id'], '—')
        ws['A6'] = "Причина:"
        ws['B6'] = act['reason'] or '—'

        row = 8
        for col, h in enumerate(["№", "Наименование", "Артикул", "Кол-во", "Цена", "Сумма"], 1):
            cell = ws.cell(row=row, column=col, value=h)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E5E7EB", end_color="E5E7EB", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')

        for idx, item in enumerate(items, 1):
            row += 1
            ws.cell(row=row, column=1, value=idx)
            ws.cell(row=row, column=2, value=item['part_name'])
            ws.cell(row=row, column=3, value=item['article'])
            ws.cell(row=row, column=4, value=f"{item['quantity']:.2f} {item['unit']}")
            ws.cell(row=row, column=5, value=item['price'])
            ws.cell(row=row, column=6, value=item['amount'])

        row += 1
        ws.merge_cells(f'A{row}:E{row}')
        tc = ws.cell(row=row, column=1, value="ИТОГО:")
        tc.font = Font(bold=True)
        tc.alignment = Alignment(horizontal='right')
        ws.cell(row=row, column=6, value=act['total_amount']).font = Font(bold=True)

        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 15

        wb.save(file_path)
        QMessageBox.information(self, "Успех", f"Файл сохранён: {file_path}")

    def _export_selected(self):
        act_id = self._get_selected_act_id()
        if not act_id:
            QMessageBox.information(self, "Подсказка", "Выберите акт для экспорта")
            return

        acts = self._load_acts()
        act = next((a for a in acts if a['id'] == act_id), None)
        if not act:
            return

        items = self._load_act_items(act_id)
        self._export_act_to_excel(act, items)

    def _export_all(self):
        acts = self._load_acts()
        if not acts:
            QMessageBox.information(self, "Подсказка", "Нет актов для экспорта")
            return

        file_path, _ = QFileDialog.getSaveFileName(
            self, "Экспорт всех актов",
            f"Акты_списания_{datetime.now().strftime('%Y%m%d')}.xlsx",
            "Excel Files (*.xlsx)"
        )
        if not file_path:
            return

        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill

            wb = openpyxl.Workbook()
            wb.remove(wb.active)

            drivers = {d.id: d.fio for d in db.list_drivers(self.conn)}
            vehicles = {v.id: f"{v.marka} {v.model}" for v in db.list_vehicles(self.conn)}

            for act in acts:
                sheet_name = act['act_number'][:31]
                ws = wb.create_sheet(title=sheet_name)

                ws.merge_cells('A1:F1')
                ws['A1'].value = f"АКТ СПИСАНИЯ № {act['act_number']}"
                ws['A1'].font = Font(size=14, bold=True)
                ws['A1'].alignment = Alignment(horizontal='center')

                ws['A3'] = "Дата:"
                ws['B3'] = act['act_date']
                ws['A4'] = "Автомобиль:"
                ws['B4'] = vehicles.get(act['vehicle_id'], '?')
                ws['A5'] = "Водитель:"
                ws['B5'] = drivers.get(act['driver_id'], '?')
                ws['A6'] = "Причина:"
                ws['B6'] = act['reason'] or ''

                row = 8
                for col, h in enumerate(["№", "Наименование", "Артикул", "Кол-во", "Цена", "Сумма"], 1):
                    c = ws.cell(row=row, column=col, value=h)
                    c.font = Font(bold=True)
                    c.fill = PatternFill(start_color="E5E7EB", end_color="E5E7EB", fill_type="solid")

                items = self._load_act_items(act['id'])
                for idx, item in enumerate(items, 1):
                    row += 1
                    ws.cell(row=row, column=1, value=idx)
                    ws.cell(row=row, column=2, value=item['part_name'])
                    ws.cell(row=row, column=3, value=item['article'])
                    ws.cell(row=row, column=4, value=f"{item['quantity']:.2f} {item['unit']}")
                    ws.cell(row=row, column=5, value=item['price'])
                    ws.cell(row=row, column=6, value=item['amount'])

                ws.column_dimensions['A'].width = 5
                ws.column_dimensions['B'].width = 30
                ws.column_dimensions['C'].width = 15
                ws.column_dimensions['D'].width = 15
                ws.column_dimensions['E'].width = 12
                ws.column_dimensions['F'].width = 15

            wb.save(file_path)
            QMessageBox.information(self, "Успех", f"Экспортировано {len(acts)} актов: {file_path}")

        except ImportError:
            QMessageBox.warning(self, "Ошибка", "openpyxl не установлен")
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", str(e))

    def _export_act_to_excel(self, act: dict, items: list):
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
        except ImportError:
            QMessageBox.warning(self, "Ошибка", "openpyxl не установлен")
            return

        drivers = {d.id: d.fio for d in db.list_drivers(self.conn)}
        vehicles = {v.id: f"{v.marka} {v.model} ({v.gosnomer})" for v in db.list_vehicles(self.conn)}

        file_path, _ = QFileDialog.getSaveFileName(
            self, "Сохранить акт",
            f"Акт_{act['act_number']}_{act['act_date']}.xlsx",
            "Excel Files (*.xlsx)"
        )
        if not file_path:
            return

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Акт списания"

            ws.merge_cells('A1:F1')
            ws['A1'].value = f"АКТ СПИСАНИЯ № {act['act_number']}"
            ws['A1'].font = Font(size=14, bold=True)
            ws['A1'].alignment = Alignment(horizontal='center')

            ws['A3'] = "Дата:"
            ws['B3'] = act['act_date']
            ws['A4'] = "Автомобиль:"
            ws['B4'] = vehicles.get(act['vehicle_id'], '?')
            ws['A5'] = "Водитель:"
            ws['B5'] = drivers.get(act['driver_id'], '?')
            ws['A6'] = "Причина:"
            ws['B6'] = act['reason'] or ''

            row = 8
            for col, h in enumerate(["№", "Наименование", "Артикул", "Кол-во", "Цена", "Сумма"], 1):
                c = ws.cell(row=row, column=col, value=h)
                c.font = Font(bold=True)
                c.fill = PatternFill(start_color="E5E7EB", end_color="E5E7EB", fill_type="solid")

            for idx, item in enumerate(items, 1):
                row += 1
                ws.cell(row=row, column=1, value=idx)
                ws.cell(row=row, column=2, value=item['part_name'])
                ws.cell(row=row, column=3, value=item['article'])
                ws.cell(row=row, column=4, value=f"{item['quantity']:.2f} {item['unit']}")
                ws.cell(row=row, column=5, value=item['price'])
                ws.cell(row=row, column=6, value=item['amount'])

            row += 1
            ws.merge_cells(f'A{row}:E{row}')
            ws.cell(row=row, column=1, value="ИТОГО:").font = Font(bold=True)
            ws.cell(row=row, column=6, value=act['total_amount']).font = Font(bold=True)

            row += 3
            ws.cell(row=row, column=1, value="Списал: ____________________")
            ws.cell(row=row, column=4, value="Принял: ____________________")

            ws.column_dimensions['A'].width = 5
            ws.column_dimensions['B'].width = 30
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 15
            ws.column_dimensions['E'].width = 12
            ws.column_dimensions['F'].width = 15

            wb.save(file_path)
            QMessageBox.information(self, "Успех", f"Файл сохранён: {file_path}")

        except Exception as e:
            QMessageBox.critical(self, "Ошибка", str(e))

    def _delete_selected(self):
        act_id = self._get_selected_act_id()
        if not act_id:
            QMessageBox.information(self, "Подсказка", "Выберите акт для удаления")
            return

        reply = QMessageBox.question(
            self, "Подтверждение", "Удалить выбранный акт списания?",
            QMessageBox.Yes | QMessageBox.No
        )
        if reply == QMessageBox.Yes:
            try:
                # Soft-delete: помечаем как удалённый
                from zavgar_app import db
                db.soft_delete(self.conn, 'write_offs', act_id)
                self._refresh()
            except Exception as e:
                QMessageBox.critical(self, "Ошибка", str(e))
